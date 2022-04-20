from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
import msoffcrypto, io, os, re, csv

# Todo: 
# extract_data() -- detect substrings (if sheetname is  [SHA1 (v2)], need to match SHA1)

#############################################################################################

# Common Sheet Names
v1 = [ "MD5", "SHA", "SHA1", "SHA256", "SHA512"]
v2 = ["IP", "Maicious_Domain(s)_IP", "DOMAIN", "URL" , "HOSTNAME"]

# store classifed sheet names 
sheet_name_hashes = []
sheet_name_address = []
sheet_name_unknown = []

# holding list for extracted data
holding_list_hash = []
holding_list_address = []
holding_list_sheet_known = []

# Blacklist Output- MD5, SHA1, SHA256, SHA512, Attacker IP, Target IP, URL
sanitized_words = {"hxxp://":"http://", "hxxps://":"https://" }
hash_types = ["md5", "sha1", "sha256", "sha512"]
output_classified = {"md5":[], "sha1":[], "sha256":[], "sha512":[], "ip":[], "url":[]}
output_unknown = {"hash":[], "ip/url":[], "sheet":[]}

#############################################################################################

# check if file is password protected
def isExcelEncrypted(xd):
    try:
        fileHandle = open(xd, "rb")
        ofile = msoffcrypto.OfficeFile(fileHandle)
        isEncrypted = ofile.is_encrypted()
        fileHandle.close()
        return isEncrypted
    except Exception as err:
        return "Exception: "+ str( format(err) )


# validate ipv4 format xxx.xxx.xxx.xxx (each xxx between 0 and 255)
def validate_ipv4(xd):
    a = xd.split('.')
    if len(a) != 4:
        # print (str(a) + " does not have 4 parts of an IP address")
        return False
    for x in a:
        if not x.isdigit():
            # print (str(a) + " does not consist of only numbers like an IP address")
            return False
        i = int(x)
        if i < 0 or i > 255:
            # print (str(a) + " range is not between 0 and 255, invalid IP")
            return False
    return True

def desanitize_url(xd):
    for word in sanitized_words:
        if word in xd:
            new_xd = xd.replace(word, sanitized_words[word])
            print ("unsanitized {} --> {}" .format(xd, new_xd))
            return new_xd
        else:
            return xd

# extract sheetnames and save to list
def get_sheet_names(xd):
    wb2 = load_workbook(xd)
    sheet_name_list = wb2.sheetnames
    print ("\nSheet Names Found ---> {} " .format(sheet_name_list))
    return sheet_name_list


# extract data from first column of verified sheet names into a holding list
def extract_data(xd):
    for i in xd:
        if i in v1:
            sheet_name_hashes.append(i)
        if i in v2:
            sheet_name_address.append(i)
        if (i not in v1) and (i not in v2):
            sheet_name_unknown.append(i)
    
    # extract data from identified sheets that have hashes/ips/urls to a holding list
    if len(sheet_name_hashes) > 0:
        for x in sheet_name_hashes:
            df = pd.read_excel(file_name, x)
            first_column = df.iloc[:, 0].tolist()
            holding_list_hash.extend(first_column)
    if len(sheet_name_address) > 0:
        for x in sheet_name_address:
            df = pd.read_excel(file_name, x)
            first_column = df.iloc[:, 0].tolist()
            holding_list_address.extend(first_column)
    # extract data from unknown sheets to a holding list
    if len(sheet_name_unknown) > 0:
        for x in sheet_name_unknown:
            df = pd.read_excel(file_name, x)
            first_column = df.iloc[:, 0].tolist()
            holding_list_sheet_known.extend(first_column)

    # get count of each holding list
    holding_list_hash_count = len(holding_list_hash)
    holding_list_address_count = len(holding_list_address)
    holding_list_sheet_known_count = len(holding_list_sheet_known)

    if holding_list_hash_count > 0:
        print ("Extracted {} potential hashes from Sheet(s) {} " .format(holding_list_hash_count, sheet_name_hashes))
    if holding_list_address_count > 0:
        print ("Extracted {} potential ip/urls from Sheet(s) {} " .format(holding_list_address_count, sheet_name_address))
    if holding_list_sheet_known_count > 0:
        print ("Extracted {} data from Unknown Sheet(s) {} " .format(holding_list_sheet_known_count, sheet_name_unknown))
    # return count of each holding list for next function (process_data)
    return [holding_list_hash_count, holding_list_address_count, holding_list_sheet_known_count]


# classify data from the holding list into various lists using regex (md5, sha1, ip, url etc)
def process_data(xd):
    # check if holding_list_hash is not empty
    if xd[0] > 0:
        # remove any whitespaces in the list 
        clean_holding_list_hash = [w.strip() for w in holding_list_hash]
        # regex for various hashes
        for potential_hash in clean_holding_list_hash:
            if re.match(r"\b([a-f\d]{32}|[A-F\d]{32})\b", potential_hash):
                output_classified["md5"].append(potential_hash)
            elif re.match(r"\b([a-f\d]{40}|[A-F\d]{40})\b", potential_hash):
                output_classified["sha1"].append(potential_hash)
            elif re.match(r"\b([a-f\d]{64}|[A-F\d]{64})\b", potential_hash):
                output_classified["sha256"].append(potential_hash)
            elif re.match(r"\b([a-f\d]{128}|[A-F\d]{128})\b", potential_hash):
                output_classified["sha512"].append(potential_hash)
            else:
                output_unknown["hash"].append(potential_hash)

    # check if holding_list_address is not empty
    if xd[1] > 0:
        # replace any [.] with . in the list
        clean_holding_list_address = [w.replace("[.]", ".") for w in holding_list_address]
        # remove any whitespaces in the list 
        cleaner_holding_list_address = [w.strip() for w in clean_holding_list_address]
        for i in cleaner_holding_list_address:
            # only add entries with . inside to remove invalid data like words and headers  
            if "." not in i:
                output_unknown["ip/url"].append(i)
            # run ipv4 validate function
            elif validate_ipv4(i) is True:
                output_classified["ip"].append(i)
            # most likely a URL if not ipv4 [Needs testing] <-------------------------------------------------------------------------
            else:
                unsanitized_i = desanitize_url(i)
                output_classified["url"].append(unsanitized_i)

    # check if holding_list_unknown is not empty
    if xd[2] > 0:
        output_unknown["sheet"].extend(holding_list_sheet_known)

    # remove dictionary keys with empty list - ioc types that don't have anything
    for i in output_classified.copy():
        if not output_classified[i]:
            output_classified.pop(i)

    for i in output_unknown.copy():
        if not output_unknown[i]:
            output_unknown.pop(i)

    # print summary of classified data
    if len(output_classified) > 0:
        print ("\nClassified Data Count: ")
        for x in output_classified:
            print ("{} --> {}" .format(x, len(output_classified[x])))
            # print (output_classified[x])
    else:
        print ("\nERROR: No extracted data was classified")

    # print summary of unknown data
    if len(output_unknown) > 0:       
        print ("\nUnknown Data Count:")
        for x in output_unknown:
            print ("{} --> {}" .format(x, len(output_unknown[x])))
            # print (output_unknown[x])
    else:
        print ("\nNo Unknown Data")

    # return count of classified and unknown data for next function (csv_generate)
    return [len(output_classified), len(output_unknown)]


# Create csv files for classified and unknown data
def csv_generate(xd):
    dtnow = datetime.now()
    dt_string = dtnow.strftime("%d-%m-%Y, %H%M%S")
    folder_string = "auto-ioc-output ({})" .format(dt_string)
    os.makedirs(folder_string)
    title_csv = "/auto-ioc-"

    print ("\n")

    if xd[0] > 0:
    # loop through list of classified ioc types
        for ioc_type in output_classified:
            # add header to csv if ioc type is a hash
            if ioc_type in hash_types:
                with open (folder_string + title_csv + ioc_type + ".csv", 'w', newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow([ioc_type, "File Name"])
                    for data in output_classified[ioc_type]:
                        writer.writerow([data])
                    print ("Generated auto-ioc-{}.csv" .format(ioc_type))
            # if not hash ioc, no header required          
            else:
                with open (folder_string + title_csv + ioc_type + ".csv", 'w', newline="") as f:
                    writer = csv.writer(f)
                    for data in output_classified[ioc_type]:
                        writer.writerow([data])
                    print ("Generated auto-ioc-{}.csv" .format(ioc_type))

    if xd[1] > 0:
        # Create csv for unknown data and dump all entries into it 
        with open (folder_string + title_csv + "unknown" + ".csv", 'w', newline="") as f:
                writer = csv.writer(f)
                for source in output_unknown:
                    for data in output_unknown[source]:
                        writer.writerow([data])
        print ("Generated auto-ioc-unknown.csv")

    print ("\n")   

#############################################################################################

# Running Code 
if __name__ == "__main__":
    # Finds .xlsx file in directory
    file_count = 0
    for file in os.listdir("."):
        if file.endswith(".xlsx"):
            file_name = file
            file_count += 1

    # If multiple .xslx file found, quit
    if file_count > 1:
        print ("\nERROR: Multiple xlsx files found, only one file can be processed, remove the rest and rerun the script")
        quit()

    # If no file found, quit
    if 'file_name' not in globals():
        print ("\nERROR: No xlsx file found")
        quit()

    # Ask for password if encrypted
    if isExcelEncrypted(file_name) is True:
        temp = io.BytesIO()
        with open (file_name, 'rb') as f:
            excel = msoffcrypto.OfficeFile(f)
            excel.load_key(input("\n{} is encrypted, enter password: " .format(file_name)))
            excel.decrypt(temp)
            file_name = temp

            sheet_names = get_sheet_names(file_name)
            holding_list_counts = extract_data(sheet_names)
            ioc_counts = process_data(holding_list_counts)
            csv_generate(ioc_counts)
            print ("auto-ioc process has completed")
    else:
        print ("\n{} is not encrypted, no password required" .format(file_name))
        sheet_names = get_sheet_names(file_name)
        holding_list_counts = extract_data(sheet_names)
        ioc_counts = process_data(holding_list_counts)
        csv_generate(ioc_counts)
        print ("auto-ioc process has completed")