from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
import msoffcrypto
import io
import os
import re
import csv
import shutil
import sys
from esm_funcs import * 


"""
todo:
more fields (source, info, cve etc)
"""

#############################################################################################

# store classifed sheet names 
sheet_name_hashes = []
sheet_name_address = []
sheet_name_unknown = []
empty_sheets_list = []

# holding list for extracted data
holding_list_hash = []
holding_list_address = []
holding_list_sheet_unknown = []

# Blacklist Output- MD5, SHA1, SHA256, SHA512, Attacker IP, Target IP, URL
hash_types = ["MD5", "SHA1", "SHA256", "SHA512"]
output_classified = {"MD5":[], "SHA1":[], "SHA256":[], "SHA512":[], "IP":[], "URL":[]}
output_unknown = {"hash":[], "ip/url":[], "sheet":[]}


#############################################################################################

# load list of recognised sheet names from sheet_address.txt and sheet_hash.txt
def load_sheet_names(xd):
    try:
        with open (xd, "r") as f:
            reader = [w.strip() for w in f.readlines()]
            print ("Recognised Sheet Names {}" .format(str(reader)))
            return reader
    except Exception as e:
        print ("Error ---> {}".format(e))
        input('Press Enter to Exit...')

# check if file is password protected
def isExcelEncrypted(xd):
    try:
        fileHandle = open(xd, "rb")
        ofile = msoffcrypto.OfficeFile(fileHandle)
        isEncrypted = ofile.is_encrypted()
        fileHandle.close()
        return isEncrypted
    except Exception as err:
        return "Exception: "+ str( format(err) )

# validate ipv4 format xxx.xxx.xxx.xxx (each xxx between 0 and 255)
def validate_ipv4(xd):
    a = xd.split('.')
    if len(a) != 4:
        # print (str(a) + " does not have 4 parts of an IP address")
        return False
    for x in a:
        if not x.isdigit():
            # print (str(a) + " does not consist of only numbers like an IP address")
            return False
        i = int(x)
        if i < 0 or i > 255:
            # print (str(a) + " range is not between 0 and 255, invalid IP")
            return False
    return True

# if hxxp / hxxps in url, replace. else return original
# additional desantise for urls with brackets like [http], [.com], [http://]
def desanitize_url(xd):
    old_xd = xd
    count = 0
    to_sanitize = {}
    sanitized_words = {"hxxp://":"http://", "hxxps://":"https://"}
    bracket_regex = re.compile(r"\[[a-zA-Z0-9\:\/\.\_\\]{1,}\]")
    matches = bracket_regex.findall(xd)
    
    for i in matches:
        to_sanitize[i]=i[1:-1]

    for word in to_sanitize:
        if word in xd:
            xd = xd.replace(word, to_sanitize[word])
            count += 1

    for word in sanitized_words:
        if word in xd:
            xd = xd.replace(word, sanitized_words[word])
            count += 1

    if count > 0:
        print ("[DESANITIZE URL] --- {} ---> {}" .format(old_xd, xd))
        return xd
    return xd

# run multiple 'cleaning' functions
def clean_list(xd):
        xd = list(map(str, xd))
        # replace [.] and [:] 
        xd1 = [w.replace("[.]", ".") for w in xd]
        xd2 = [w.replace("[:]", ":") for w in xd1]
        # remove any whitespaces in the list 
        xd3 = [w.strip() for w in xd2]
        return xd3

# search for port numbers and remove them
def sanitize_ports(xd):
    new_xd = []
    for string in xd:
        check = ([(m.span()) for m in re.finditer(r"\b:[\d]{1,}\b", string)])
        if len(check) > 0:
            indexes = check[0]
            impt_index = indexes[0]
            new_string = string[:impt_index]
            print ("[SANITIZE PORTS] --- {} ---> {}" .format(string, new_string))
            new_xd.append(new_string)
        else:
            new_xd.append(string)
    return new_xd

# extract sheetnames and save to list
def get_sheet_names(xd):
    wb2 = load_workbook(xd)
    sheet_name_list = wb2.sheetnames
    print ("\nSheet Names Found ---> {} " .format(sheet_name_list))
    return sheet_name_list


# extract data from first column of verified sheet names into a holding list
def extract_data(xd, sh_hash, sh_address):
    for i in xd:
        if i.upper() in sh_hash:
            sheet_name_hashes.append(i) 
        if i.upper() in sh_address:
            sheet_name_address.append(i)
        if (i.upper() not in sh_hash) and (i.upper() not in sh_address):
            sheet_name_unknown.append(i)
    
    # extract data from identified sheets that have hashes/ips/urls to holding lists
    if len(sheet_name_hashes) > 0:
        for x in sheet_name_hashes:
            df = pd.read_excel(file_name, x, header=None)
            is_empty = df.empty
            if not is_empty:
                first_column = df.iloc[:, 0].tolist()
                holding_list_hash.extend(first_column)
            else:
                empty_sheets_list.append(x)
    if len(sheet_name_address) > 0:
        for x in sheet_name_address:
            df = pd.read_excel(file_name, x, header=None)
            is_empty = df.empty
            if not is_empty:
                first_column = df.iloc[:, 0].tolist()
                holding_list_address.extend(first_column)
            else:
                empty_sheets_list.append(x)
    # extract data from unknown sheets to a holding list
    if len(sheet_name_unknown) > 0:
        for x in sheet_name_unknown:
            df = pd.read_excel(file_name, x, header=None)
            is_empty = df.empty
            if not is_empty:
                first_column = df.iloc[:, 0].tolist()
                holding_list_sheet_unknown.extend(first_column)
            else:
                empty_sheets_list.append(x)

    # get count of each holding list
    holding_list_hash_count = len(holding_list_hash)
    holding_list_address_count = len(holding_list_address)
    holding_list_sheet_unknown_count = len(holding_list_sheet_unknown)

    if holding_list_hash_count > 0:
        print ("Extracted {} Hashes from {} " .format(holding_list_hash_count, sheet_name_hashes))
    if holding_list_address_count > 0:
        print ("Extracted {} IP/URLs from {} " .format(holding_list_address_count, sheet_name_address))
    if holding_list_sheet_unknown_count > 0:
        print ("Extracted {} Data from Unrecognised Sheets {} " .format(holding_list_sheet_unknown_count, sheet_name_unknown))
    if len(empty_sheets_list) > 0:
        print ("Sheet {} is Empty, No Data Extracted" .format(str(empty_sheets_list)))

    # return count of each holding list for next function (process_data)
    return [holding_list_hash_count, holding_list_address_count, holding_list_sheet_unknown_count]

# classify data from the holding list into various lists using regex (md5, sha1, ip, url etc)
def process_data(xd):
    # check if holding_list_hash is not empty
    if xd[0] > 0:
        # remove any whitespaces in the list 
        clean_holding_list_hash = [w.strip() for w in holding_list_hash]
        # regex for various hashes
        for potential_hash in clean_holding_list_hash:
            if re.match(r"\b([a-f\d]{32}|[A-F\d]{32})\b", potential_hash):
                output_classified["MD5"].append(potential_hash)
            elif re.match(r"\b([a-f\d]{40}|[A-F\d]{40})\b", potential_hash):
                output_classified["SHA1"].append(potential_hash)
            elif re.match(r"\b([a-f\d]{64}|[A-F\d]{64})\b", potential_hash):
                output_classified["SHA256"].append(potential_hash)
            elif re.match(r"\b([a-f\d]{128}|[A-F\d]{128})\b", potential_hash):
                output_classified["SHA512"].append(potential_hash)
            else:
                output_unknown["hash"].append(potential_hash)

    # check if holding_list_address is not empty
    if xd[1] > 0:
        print ("\n")
        clean_holding_list_address = clean_list(holding_list_address)
        clean_holding_list_address = sanitize_ports(clean_holding_list_address)

        for i in clean_holding_list_address:
            # only add entries with . inside to remove invalid data like words and headers  
            if "." not in i:
                print("[NON ADDRESS/IP] --- {} ---> Sent to List of Unknowns".format(i))
                output_unknown["ip/url"].append(i)
            # remove false positive edge cases (headers with a .)
            elif re.match(r"\B\.[a-zA-Z0-9]{1,}|[. a-zA-Z0-9]{1,}\.\B", i):
                print("[NON ADDRESS/IP] --- {} ---> Sent to List of Unknowns".format(i))
                output_unknown["ip/url"].append(i)
            # run ipv4 validate function
            elif validate_ipv4(i) is True:
                output_classified["IP"].append(i)
            # most likely a URL if not ipv4 
            else:
                unsanitized_i = desanitize_url(i)
                output_classified["URL"].append(unsanitized_i)

    # check if holding_list_unknown is not empty
    if xd[2] > 0:
        output_unknown["sheet"].extend(holding_list_sheet_unknown)

    # remove dictionary keys with empty list - ioc types that don't have anything
    for i in output_classified.copy():
        if not output_classified[i]:
            output_classified.pop(i)

    for i in output_unknown.copy():
        if not output_unknown[i]:
            output_unknown.pop(i)

    # print summary of classified data
    if len(output_classified) > 0:
        print ("\nClassified Data Count: ")
        for x in output_classified:
            print ("{} ---> {}" .format(x, len(output_classified[x])))
            # print (output_classified[x])
    else:
        print ("\nERROR: Failed to Classify Data")

    # print summary of unknown data
    if len(output_unknown) > 0:       
        print ("\nUnknown Data Count:")
        for x in output_unknown:
            print ("{} ---> {}" .format(x, len(output_unknown[x])))
            # print (output_unknown[x])
    else:
        print ("\nNo Unknown Data")

    # return count of classified and unknown data for next function (csv_generate)
    return [len(output_classified), len(output_unknown)]

# Create csv files for classified and unknown data
def csv_generate(xd, src_file, pw):
    dtnow = datetime.now()
    dt_string = dtnow.strftime("%d-%m-%Y, %H%M%S")
    folder_string = "[ouput] {}  ({})" .format(og_file_name, dt_string)
    os.makedirs(folder_string)
    title_csv = "/auto-ioc-"

    print ("\n")

    if xd[0] > 0:
    # loop through list of classified ioc types
        for ioc_type in output_classified:
            # add header to csv if ioc type is a hash, also use , as delimiter
            if ioc_type in hash_types:
                with open (folder_string + title_csv + ioc_type + ".csv", 'w', newline="") as f:
                    writer = csv.writer(f,  delimiter=",")
                    writer.writerow([ioc_type, "File Name"])
                    for data in output_classified[ioc_type]:
                        writer.writerow([data] + [""]) # add whitespace to create empty entry in 'filename' column, forcing a , delimiter
                    print ("Generated auto-ioc-{}.csv" .format(ioc_type))
            # if not hash ioc, no header required          
            else:
                with open (folder_string + title_csv + ioc_type + ".csv", 'w', newline="") as f:
                    writer = csv.writer(f)
                    for data in output_classified[ioc_type]:
                        writer.writerow([data])
                    print ("Generated auto-ioc-{}.csv" .format(ioc_type))

    if xd[1] > 0:
        print_unknown = []
        # Create csv for unknown data and dump all entries into it 
        with open (folder_string + title_csv + "unknown" + ".csv", 'w', newline="") as f:
                writer = csv.writer(f)
                for source in output_unknown:
                    for data in output_unknown[source]:
                        writer.writerow([data])
                        print_unknown.append(data)
        print ("Generated auto-ioc-unknown.csv")
        print ("auto-ioc-unknown.csv contains {}" .format(str(print_unknown)))

    # if excel file was decrypted with a password, save the password into saved-pw.txt in the output folder
    if pw is not None:
        print ("\npassword saved to saved-pw.txt in output folder")
        with open (folder_string + "/saved-pw.txt", 'w') as f:
            f.write(pw)
    # clear contents of pw.txt        
        with open ("pw.txt", "w") as fa:
            pass

    # make a copy of the excel file in the output folder
    dst_file = folder_string + "/" + src_file
    shutil.copy2(src_file, dst_file)

# prints out processed iocs for user to review, if all is good, function will return normally and next step is adding into esm
def ioc_review(xd):
    user_continue = input("\nPress Enter to Review IOCs")

    for x in xd:
        print ("\n{} Entries to be added \n" .format(x))
        for y in xd[x]:
            print (y)
        user_continue = input('\nDo The Entries Look Valid? Press Enter if Yes, Press Any Other Key and Enter if No: ')
        if user_continue == "":
            pass
        else:
            print ("Entries maybe invalid, cancelling the script")
            input('Press Enter to Exit...')
            sys.exit()
    print ('IOC Review Complete')

def start_esm_import():
    user_continue = input('\nTo start ESM Import Press Enter, Press Any Other Key and Enter to Cancel: ')
    if user_continue == "":
        return 0
    else:
        print ("ESM Import Cancelled")
        input('Press Enter to Exit...')
        sys.exit()
    
# wrap json over IOC entries for esm api request
def json_format_ioc_hash(xd):
    entries = """"""
    for i in xd:
        entries += """\n{"entry":['""" + i + """', '']},"""
    return entries

# wrap json over IOC entries for esm api request
def json_format_ioc_address(xd):
    entries = """"""
    for i in xd:
        entries += """\n{"entry":['""" + i + """']},"""
    return entries

# convert list of single dics (that esm returns in getActiveList) into list 
def flatten_address(xd):
    new_xd = []
    for i in xd:
        for j in i:
            new_xd.append(i[j])
    return new_xd

# convert list of single dics (that esm returns in getActiveList) into list 
def flatten_hashes(xd):
    new_xd = []
    for i in xd:
        for j in i:
            new_xd.append(i[j][0])
    return new_xd

# check if ioc entries were added to esm (compare against results from getActivelist)
def verify_ioc_added(verify_ioc, entry_list, ioc_type):
    failed_to_add = []
    add_count = 0
    full_count = len(verify_ioc)
    for i in verify_ioc:
        if i in entry_list:
            add_count += 1
        else:
            failed_to_add.append(i)
    if add_count == full_count:
        print ("Successfully Added {} IOCs to {} ActiveList" .format(full_count, ioc_type))
    else:
        print (" {} / {} IOCs were Added to {}" .format(add_count, full_count, ioc_type))
        print ("Failed to Add ---> {}" .format(str(failed_to_add)))

#############################################################################################

# Running Code 
if __name__ == "__main__":
    print ("\nRunning auto-ioc\n")
    # Finds .xlsx file in directory
    file_count = 0
    for file in os.listdir("."):
        if file.endswith(".xlsx"):
            og_file_name = file
            file_name = file
            file_count += 1

    # If multiple .xslx file found, quit
    if file_count > 1:
        print ("\nERROR: Multiple xlsx files found, only one file can be processed, remove the rest and try again")
        input('Press Enter to Exit...')
        sys.exit()

    # If no file found, quit
    if 'file_name' not in globals():
        print ("\nERROR: No xlsx file found")
        input('Press Enter to Exit...')
        sys.exit()

    default_list_hash = load_sheet_names("dependancies/sheet_hash.txt")
    default_list_address = load_sheet_names("dependancies/sheet_address.txt")

    esm_creds = load_esm_creds("dependancies/esm_credentials.txt")
    esm_hostnames = load_esm_hostnames("dependancies/esm_hostnames.txt")
    esm_resource_ids = load_resource_ids("dependancies/esm_resource_ids.txt")
    esm_resource_ids["IP"] = esm_resource_ids["IP"].split(",")

    # Require password if encrypted
    if isExcelEncrypted(file_name) is True:
        try:
            temp = io.BytesIO()
            with open (file_name, "rb") as f:
                excel = msoffcrypto.OfficeFile(f)
                with open ("pw.txt", "r") as f:
                    excel_pw = f.read()
                    excel_pw = excel_pw.strip()
                    if excel_pw == "":
                        excel_pw = input("\n{} is encrypted, pw.txt is empty, enter password: " .format(file_name))
                        excel.load_key(excel_pw)
                        excel.decrypt(temp)
                        file_name = temp
                    else:
                        print ("\n {} is encrypted, using password from pw.txt" .format(og_file_name))                    
                        excel.load_key(excel_pw)
                        excel.decrypt(temp)
                        file_name = temp

                sheet_names = get_sheet_names(file_name)
                holding_list_counts = extract_data(sheet_names, default_list_hash, default_list_address)
                ioc_counts = process_data(holding_list_counts)
                csv_generate(ioc_counts, og_file_name, excel_pw)
                ioc_review(output_classified)

            """
            ESM API Code
            """
            start_esm_import()

            for esm_name in esm_hostnames:
                print ("\nLogin to {} with user {}".format(esm_name, esm_creds[0]) )
                esm_auth_token = get_auth_token(esm_creds[0], esm_creds[1], esm_name)
                if esm_auth_token:
                    print ("Login Successful --- {}" .format(esm_auth_token))
                    
                    for ioc_type in output_classified:

                        if ioc_type in hash_types:
                            json_ioc_entries_hash = json_format_ioc_hash(output_classified[ioc_type])
                            add_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type], ioc_type, json_ioc_entries_hash)
                            entry_full_list = get_activelist_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type])
                            entry_full_list_clean = flatten_hashes(entry_full_list)
                            # print ("total entry count = {}" .format(len(entry_full_list_clean)))
                            # verify_ioc_added(output_classified[ioc_type], entry_full_list_clean, ioc_type)

                        elif ioc_type == "URL":
                            json_ioc_entries_address = json_format_ioc_address(output_classified[ioc_type])
                            add_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type], "RequestUrl", json_ioc_entries_address)
                            entry_full_list = get_activelist_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type])
                            entry_full_list_clean = flatten_address(entry_full_list)
                            # print ("total entry count = {}" .format(len(entry_full_list_clean)))
                            # verify_ioc_added(output_classified[ioc_type], entry_full_list_clean, ioc_type)
                            
                        elif ioc_type == "IP":
                            json_ioc_entries_address = json_format_ioc_address(output_classified[ioc_type])

                            add_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type][0], "AttackerAddress", json_ioc_entries_address)
                            entry_full_list = get_activelist_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type][0])
                            entry_full_list_clean = flatten_address(entry_full_list)
                            # print ("total entry count = {}" .format(len(entry_full_list_clean)))
                            # verify_ioc_added(output_classified[ioc_type], entry_full_list_clean, ioc_type)

                            add_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type][1], "TargetAddress", json_ioc_entries_address)
                            entry_full_list = get_activelist_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type][1])
                            entry_full_list_clean = flatten_address(entry_full_list)
                            # print ("total entry count = {}" .format(len(entry_full_list_clean)))
                            # verify_ioc_added(output_classified[ioc_type], entry_full_list_clean, ioc_type)
                
                logout(esm_name, esm_auth_token)
                input("\nImport to {} Complete, Press Enter to Continue" .format(esm_name))

            print ("\nESM Import has Completed\n")
            input('auto-ioc has Ended, Press Enter to Exit...')
        except Exception as e:
            print ("Error ---> {}".format(e))
            input('Press Enter to Exit...')

    else:
        try:
            print ("\n{} is Not Encrypted" .format(file_name))
            sheet_names = get_sheet_names(file_name)
            holding_list_counts = extract_data(sheet_names, default_list_hash, default_list_address)
            ioc_counts = process_data(holding_list_counts)
            csv_generate(ioc_counts, og_file_name, None)
            ioc_review(output_classified)

            """
            ESM API Code
            """
            start_esm_import()

            for esm_name in esm_hostnames:
                print ("\nLogin to {} with user {}".format(esm_name, esm_creds[0]) )
                esm_auth_token = get_auth_token(esm_creds[0], esm_creds[1], esm_name)
                if esm_auth_token:
                    print ("Login Successful --- {}" .format(esm_auth_token))
                    
                    for ioc_type in output_classified:

                        if ioc_type in hash_types:
                            json_ioc_entries_hash = json_format_ioc_hash(output_classified[ioc_type])
                            add_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type], ioc_type, json_ioc_entries_hash)
                            entry_full_list = get_activelist_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type])
                            entry_full_list_clean = flatten_hashes(entry_full_list)
                            # print ("total entry count = {}" .format(len(entry_full_list_clean)))
                            # verify_ioc_added(output_classified[ioc_type], entry_full_list_clean, ioc_type)

                        elif ioc_type == "URL":
                            json_ioc_entries_address = json_format_ioc_address(output_classified[ioc_type])
                            add_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type], "RequestUrl", json_ioc_entries_address)
                            entry_full_list = get_activelist_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type])
                            entry_full_list_clean = flatten_address(entry_full_list)
                            # print ("total entry count = {}" .format(len(entry_full_list_clean)))
                            # verify_ioc_added(output_classified[ioc_type], entry_full_list_clean, ioc_type)
                            
                        elif ioc_type == "IP":
                            json_ioc_entries_address = json_format_ioc_address(output_classified[ioc_type])

                            add_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type][0], "AttackerAddress", json_ioc_entries_address)
                            entry_full_list = get_activelist_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type][0])
                            entry_full_list_clean = flatten_address(entry_full_list)
                            # print ("total entry count = {}" .format(len(entry_full_list_clean)))
                            # verify_ioc_added(output_classified[ioc_type], entry_full_list_clean, ioc_type)

                            add_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type][1], "TargetAddress", json_ioc_entries_address)
                            entry_full_list = get_activelist_entries(esm_name, esm_auth_token, esm_resource_ids[ioc_type][1])
                            entry_full_list_clean = flatten_address(entry_full_list)
                            # print ("total entry count = {}" .format(len(entry_full_list_clean)))
                            # verify_ioc_added(output_classified[ioc_type], entry_full_list_clean, ioc_type)
                
                logout(esm_name, esm_auth_token)
                input("\nImport to {} Complete, Press Enter to Continue" .format(esm_name))    
            
            print ("\nESM Import has Completed\n")
            input('auto-ioc has Ended, Press Enter to Exit...')
        except Exception as e:
            print ("Error ---> {}".format(e))
            input('Press Enter to Exit...')